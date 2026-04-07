from datetime import datetime, timedelta
import csv
import io
import json
import os
import sqlite3
import sys
import shutil
from pathlib import Path
from copy import copy
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from flask import Flask, flash, g, redirect, render_template, request, url_for, Response, send_file, send_from_directory


SOURCE_DIR = Path(__file__).resolve().parent
if getattr(sys, "frozen", False):
    RUNTIME_DIR = Path(sys.executable).resolve().parent
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", RUNTIME_DIR))
else:
    RUNTIME_DIR = SOURCE_DIR
    RESOURCE_DIR = SOURCE_DIR

if getattr(sys, "frozen", False):
    DEFAULT_DATA_DIR = Path(os.environ.get("LOCALAPPDATA", str(RUNTIME_DIR))) / "PCR_BATO"
else:
    DEFAULT_DATA_DIR = SOURCE_DIR / "instance"

APP_DATA_DIR = Path(os.environ.get("PCR_BATO_DATA_DIR", str(DEFAULT_DATA_DIR))).resolve()
DATABASE = Path(os.environ.get("PCR_DB_PATH", str(APP_DATA_DIR / "pcr.db"))).resolve()
BACKUP_DIR = APP_DATA_DIR / "backups"
TEMPLATE_XLSX = APP_DATA_DIR / "logsheet_template.xlsx"
BUNDLED_TEMPLATE_XLSX = RESOURCE_DIR / "instance" / "logsheet_template.xlsx"
XLSX_EXPORT_HEADER_ROW = 8
XLSX_EXPORT_FIRST_DATA_ROW = 9
XLSX_EXPORT_ENTRIES_PER_SHEET = 25
XLSX_EXPORT_ROW_HEIGHT = 30
XLSX_EXPORT_HEADER_FONT_SIZE = 12
XLSX_EXPORT_DATA_FONT_SIZE = 11
SEED_DEMO_DATA = os.environ.get("PCR_SEED_SAMPLE_DATA", "1").strip().lower() not in {"0", "false", "no", "off"}
XLSX_EXPORT_FIELDS = [
    ("type_of_emergency", "Type of Emergency"),
    ("chief_complaint", "Chief Complaint"),
    ("patient_name", "Name of Patient"),
    ("patient_address", "Address of Patient"),
    ("sex", "Sex"),
    ("date_of_incident", "Date of Incident"),
    ("time_of_incident", "Time of Incident"),
    ("place_of_incident", "Place of Incident"),
    ("driver", "Driver"),
    ("responders", "Responders"),
    ("communicator", "Communicator"),
    ("remarks", "Remarks"),
]

app = Flask(
    __name__,
    template_folder=str(RESOURCE_DIR / "templates"),
    static_folder=str(RESOURCE_DIR / "static"),
)
app.config["SECRET_KEY"] = "change-this-secret-key"


@app.after_request
def disable_dynamic_page_cache(response):
    # Prevent stale dashboards/forms when opened through browsers with caching/PWA.
    if response.mimetype == "text/html":
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        # Durability-oriented defaults for local SQLite usage.
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA journal_mode = WAL")
        g.db.execute("PRAGMA synchronous = FULL")
    return g.db


def init_db() -> None:
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    DATABASE.parent.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    if not TEMPLATE_XLSX.exists() and BUNDLED_TEMPLATE_XLSX.exists():
        shutil.copy2(BUNDLED_TEMPLATE_XLSX, TEMPLATE_XLSX)


def _get_template_xlsx_path() -> Path:
    if TEMPLATE_XLSX.exists():
        return TEMPLATE_XLSX
    return BUNDLED_TEMPLATE_XLSX

    db = sqlite3.connect(DATABASE)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS pcr_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_name TEXT NOT NULL,
            nature_of_call TEXT,
            call_date TEXT,
            time_of_call TEXT,
            status TEXT NOT NULL DEFAULT 'SUBMITTED',
            form_data TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    # Backward-compatible migration from the earlier simple schema.
    cols = {row[1] for row in db.execute("PRAGMA table_info(pcr_reports)").fetchall()}
    if "status" not in cols:
        db.execute("ALTER TABLE pcr_reports ADD COLUMN status TEXT NOT NULL DEFAULT 'SUBMITTED'")
    if "form_data" not in cols:
        db.execute("ALTER TABLE pcr_reports ADD COLUMN form_data TEXT NOT NULL DEFAULT '{}' ")
    db.commit()

    total_records = db.execute("SELECT COUNT(*) AS total FROM pcr_reports").fetchone()["total"]
    if total_records == 0 and SEED_DEMO_DATA:
        _seed_demo_records(db)
    db.close()


def _build_seed_record(index: int) -> tuple[str, str, str, str, str, dict]:
    first_names = [
        "Maria",
        "Jose",
        "Ana",
        "John",
        "Catherine",
        "Michael",
        "Grace",
        "Daniel",
        "Angela",
        "Mark",
    ]
    last_names = [
        "Santos",
        "Reyes",
        "Cruz",
        "Garcia",
        "Lopez",
        "Ramos",
        "Torres",
        "Navarro",
        "Bautista",
        "Flores",
    ]
    locations = [
        "Poblacion North, Bato",
        "Poblacion South, Bato",
        "San Isidro Barangay Hall",
        "Rural Health Unit Area",
        "Coastal Road, Bato",
        "Mountain View, Bato",
    ]
    facilities = [
        "Bato District Hospital",
        "RHU Bato",
        "Catbalogan Provincial Hospital",
        "Regional Medical Center",
    ]
    emergencies = [
        "Medical",
        "Trauma",
        "OB",
        "Cardiac",
        "Respiratory",
        "Neurologic",
    ]
    complaints = [
        "Chest pain",
        "Shortness of breath",
        "Fever and weakness",
        "Abdominal pain",
        "Hypertension episode",
        "Minor vehicular injury",
        "Dizziness and vomiting",
        "Laceration on forearm",
    ]
    drivers = ["A. Dela Cruz", "R. Santos", "M. Garcia", "J. Flores"]
    crews = [
        "P. Navarro, L. Torres",
        "K. Ramos, M. Flores",
        "J. Reyes, A. Bautista",
        "C. Garcia, M. Lopez",
    ]
    addresses = [
        "Brgy. Poblacion North, Bato, Leyte",
        "Brgy. Poblacion South, Bato, Leyte",
        "Brgy. San Roque, Bato, Leyte",
        "Brgy. Sto. Nino, Bato, Leyte",
        "Brgy. Libagon, Bato, Leyte",
    ]

    patient_name = f"{first_names[index % len(first_names)]} {last_names[index % len(last_names)]}"
    nature_of_call = emergencies[index % len(emergencies)]
    chief_complaint = complaints[index % len(complaints)]
    patient_address = addresses[index % len(addresses)]
    incident_location = locations[index % len(locations)]
    driver = drivers[index % len(drivers)]
    crew = crews[index % len(crews)]
    facility = facilities[index % len(facilities)]
    sex = "Male" if index % 2 == 0 else "Female"
    age = str(19 + (index % 58))
    call_date = (datetime.now().date() - timedelta(days=99 - index)).isoformat()
    time_of_call = f"{7 + (index % 12):02d}:{(index * 5) % 60:02d}"
    narr = f"Patient assessed at {incident_location}. Stable for transport to {facility}."

    form_data = {
        "patient_info": {
            "age": age,
            "gender": sex,
            "nationality": "Filipino",
            "date": call_date,
            "time_of_call": time_of_call,
        },
        "contact_dispatch": {
            "permanent_address": patient_address,
            "contact_number": f"09{index % 10}{(index + 2) % 10}{(index + 4) % 10}{(index + 6) % 10}{(index + 8) % 10}{(index + 1) % 10}{(index + 3) % 10}{(index + 5) % 10}{(index + 7) % 10}",
            "etd_base": time_of_call,
            "types_of_emergencies": [nature_of_call],
            "eta_scene": f"{8 + (index % 18)} mins",
        },
        "next_of_kin": {
            "kins": [
                {
                    "name": f"{first_names[(index + 3) % len(first_names)]} {last_names[(index + 4) % len(last_names)]}",
                    "contact": f"09{(index + 1) % 10}{(index + 3) % 10}{(index + 5) % 10}{(index + 7) % 10}{(index + 9) % 10}{(index + 2) % 10}{(index + 4) % 10}{(index + 6) % 10}{(index + 8) % 10}{(index + 0) % 10}",
                }
            ],
        },
        "incident_details": {
            "location_of_incident": incident_location,
            "nature_of_illness": chief_complaint,
            "mechanism_of_injury": "N/A" if nature_of_call != "Trauma" else "Slip and fall",
            "etd_scene": f"{12 + (index % 10)} mins",
            "eta_hospital": f"{15 + (index % 12)} mins",
            "etd_hospital": f"{20 + (index % 10)} mins",
            "eta_base": f"{25 + (index % 12)} mins",
        },
        "patient_assessment": {
            "chief_complaint": chief_complaint,
            "c_spine": "No" if nature_of_call in {"Medical", "Cardiac", "Respiratory", "Neurologic"} else "Yes",
            "airway": "Patent",
            "breathing": ["Spontaneous"],
            "circulation": ["Radial pulse present"],
            "pupils": "PERRLA",
            "loc_level": "Alert",
            "capillary_refill": "<2 sec",
        },
        "gcs": {
            "eye_opening": "4",
            "verbal_response": "5",
            "motor_response": "6",
            "gcs_total": "15",
        },
        "vital_signs": [
            {
                "time": time_of_call,
                "loc": "Alert",
                "bp": f"{110 + (index % 20)}/{70 + (index % 10)}",
                "rr": str(16 + (index % 4)),
                "pr": str(78 + (index % 18)),
                "temp": f"{36.4 + (index % 6) * 0.1:.1f}",
                "spo2": str(97 - (index % 3)),
                "rbs": str(95 + (index % 30)),
                "pain_scale": str(index % 10),
                "gcs_eye": "4",
                "gcs_verbal": "5",
                "gcs_motor": "6",
                "gcs_total": "15",
            }
        ],
        "sample_history": {
            "symptoms": chief_complaint,
            "allergies": "None reported",
            "medications": "Maintenance meds as prescribed",
            "past_medical_history": "Hypertension" if index % 3 == 0 else "N/A",
            "last_oral_intake": "Within 4 hours",
            "events_prior": "Onset noted before EMS arrival",
        },
        "opqrst": {
            "onset": "Sudden",
            "provoking_factors": "Activity",
            "quality": "Dull",
            "radiation": "None",
            "severity": str(3 + (index % 7)),
            "timing": "Intermittent",
        },
        "physical_exam": {
            "skin": ["Warm", "Dry"],
            "burns": [],
            "findings": ["No obvious deformity"],
            "body_diagram_drawing": "",
            "deformity": "No",
            "bleeding": "No",
            "contusion": "No",
            "tenderness": "Mild",
            "abrasion": "No",
            "laceration": "No",
            "punctured": "No",
            "swelling": "No",
        },
        "narrative_report": narr,
        "team_destination": {
            "ambulance_driver": driver,
            "plate_no": f"AMB-{100 + index:03d}",
            "license_no": f"DL-{20000 + index}",
            "responders_tl": crew.split(",")[0].strip(),
            "crew": crew,
            "crew_members": [member.strip() for member in crew.split(",")],
            "destination_determination": facility,
            "receiving_facility": facility,
            "communicator": "Dispatch",
            "receiving_personnel": f"Nurse {last_names[(index + 2) % len(last_names)]}",
        },
        "care_management": {
            "airway": ["Positioning"],
            "breathing": ["Oxygen support"],
            "circulation": ["Monitoring"],
            "immobilization": ["None" if nature_of_call != "Trauma" else "Spinal precautions"],
            "wound_care": ["Dressing" if nature_of_call == "Trauma" else "None"],
        },
        "mvc": {
            "type_of_vehicles": "Car vs Motorcycle" if nature_of_call == "Trauma" else "N/A",
            "plate_numbers": f"ABC-{300 + index}",
            "type_of_accident": "Minor collision" if nature_of_call == "Trauma" else "N/A",
            "law_enforcer_name": f"Officer {last_names[(index + 5) % len(last_names)]}",
            "law_enforcer_contact": f"0917{index:07d}"[-11:],
        },
        "consent_refusal": {
            "consent_patient_name": patient_name,
            "consent_date": call_date,
            "consent_time": time_of_call,
            "refusal_treatment_agreement": "Agreed",
            "refusal_admission_agreement": "Agreed",
        },
    }

    return patient_name, nature_of_call, call_date, time_of_call, sex, form_data


def _seed_demo_records(db: sqlite3.Connection) -> None:
    for index in range(100):
        patient_name, nature_of_call, call_date, time_of_call, sex, form_data = _build_seed_record(index)
        db.execute(
            """
            INSERT INTO pcr_reports (
                patient_name, nature_of_call, call_date, time_of_call,
                status, form_data, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                patient_name,
                nature_of_call,
                call_date,
                time_of_call,
                "SUBMITTED",
                json.dumps(form_data, ensure_ascii=True),
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
    db.commit()


def _format_backup_date(value: str) -> str | None:
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None
    return f"{parsed.strftime('%B')} {parsed.day} {parsed.year}"


def _build_backup_filename() -> str:
    with sqlite3.connect(DATABASE) as db:
        db.row_factory = sqlite3.Row
        row = db.execute(
            """
            SELECT MIN(call_date) AS first_date, MAX(call_date) AS last_date
            FROM pcr_reports
            WHERE call_date IS NOT NULL AND TRIM(call_date) <> ''
            """
        ).fetchone()

    first_date = _format_backup_date(row["first_date"] or "") if row else None
    last_date = _format_backup_date(row["last_date"] or "") if row else None

    if first_date and last_date:
        return f"{first_date} - {last_date}.db"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"pcr_backup_{timestamp}.db"


def _create_db_backup_file() -> Path:
    backup_path = BACKUP_DIR / _build_backup_filename()
    if backup_path.exists():
        backup_path.unlink()

    src = get_db()
    dest = sqlite3.connect(backup_path)
    try:
        src.backup(dest)
    finally:
        dest.close()

    return backup_path


# Ensure schema exists whether app is started via `python app.py` or `flask run`.
init_db()


@app.teardown_appcontext
def close_db(_error: BaseException | None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.route("/")
def home():
    return redirect(url_for("records"))


@app.route("/manifest.webmanifest")
def webmanifest() -> Response:
    return send_from_directory(RESOURCE_DIR / "static", "manifest.webmanifest", mimetype="application/manifest+json")


@app.route("/sw.js")
def service_worker() -> Response:
    response = send_from_directory(RESOURCE_DIR / "static", "sw.js", mimetype="application/javascript")
    response.headers["Service-Worker-Allowed"] = "/"
    return response


def _collect_multi(field: str) -> list[str]:
    return [v for v in request.form.getlist(field) if v]


def _collect_kin_entries() -> list[dict[str, str]]:
    """Collect dynamically added next of kin entries."""
    kins: list[dict[str, str]] = []
    indexes = sorted(
        {
            int(key.rsplit("_", 1)[1])
            for key in request.form.keys()
            if key.startswith("kin_name_") and key.rsplit("_", 1)[1].isdigit()
        }
        | {
            int(key.rsplit("_", 1)[1])
            for key in request.form.keys()
            if key.startswith("kin_contact_") and key.rsplit("_", 1)[1].isdigit()
        }
    )
    for idx in indexes:
        name = request.form.get(f"kin_name_{idx}", "").strip()
        contact = request.form.get(f"kin_contact_{idx}", "").strip()
        if name or contact:
            kins.append({"name": name, "contact": contact})
    return kins


def _collect_vital_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for idx in range(1, 6):
        row = {
            "time": request.form.get(f"vital_time_{idx}", "").strip(),
            "loc": request.form.get(f"vital_loc_{idx}", "").strip(),
            "bp": request.form.get(f"vital_bp_{idx}", "").strip(),
            "rr": request.form.get(f"vital_rr_{idx}", "").strip(),
            "pr": request.form.get(f"vital_pr_{idx}", "").strip(),
            "temp": request.form.get(f"vital_temp_{idx}", "").strip(),
            "spo2": request.form.get(f"vital_spo2_{idx}", "").strip(),
            "rbs": request.form.get(f"vital_rbs_{idx}", "").strip(),
            "pain_scale": request.form.get(f"vital_pain_scale_{idx}", "").strip(),
            "gcs_eye": request.form.get(f"vital_gcs_eye_{idx}", "").strip(),
            "gcs_verbal": request.form.get(f"vital_gcs_verbal_{idx}", "").strip(),
            "gcs_motor": request.form.get(f"vital_gcs_motor_{idx}", "").strip(),
            "gcs_total": request.form.get(f"vital_gcs_total_{idx}", "").strip(),
        }
        if any(row.values()):
            rows.append(row)
    return rows


def _split_csv_values(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def _agreement_status(values: list[str] | str | None) -> str:
    if isinstance(values, list):
        return "Agreed" if values else "Disagreed"
    if isinstance(values, str):
        normalized = values.strip().lower()
        if normalized in {"agreed", "confirmed", "yes", "true", "1"}:
            return "Agreed"
        if normalized in {"disagreed", "no", "false", "0"}:
            return "Disagreed"
    return "Disagreed"


def _load_form_data(row: sqlite3.Row) -> dict:
    try:
        return json.loads(row["form_data"] or "{}")
    except json.JSONDecodeError:
        return {}


def _build_xlsx_workbook(rows: list[sqlite3.Row]) -> Workbook:
    template_path = _get_template_xlsx_path()
    if template_path.exists():
        workbook = load_workbook(template_path)
    else:
        workbook = Workbook()

    thin_side = Side(style="thin", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    chunks = [rows[index:index + XLSX_EXPORT_ENTRIES_PER_SHEET] for index in range(0, len(rows), XLSX_EXPORT_ENTRIES_PER_SHEET)]

    if not chunks:
        chunks = [[]]

    base_sheet = workbook.active

    for sheet_index, chunk in enumerate(chunks, start=1):
        if sheet_index == 1:
            worksheet = base_sheet
            worksheet.title = "PCR Records"
        else:
            worksheet = workbook.create_sheet(title=f"PCR Records {sheet_index}")
            # Match layout so extra sheets don't look compressed.
            for column_letter, dimension in base_sheet.column_dimensions.items():
                if dimension.width is not None:
                    worksheet.column_dimensions[column_letter].width = dimension.width
            for row_num in range(1, XLSX_EXPORT_FIRST_DATA_ROW):
                base_height = base_sheet.row_dimensions[row_num].height
                if base_height is not None:
                    worksheet.row_dimensions[row_num].height = base_height

            # Copy the visible template/header block so continuation sheets look the same.
            for row_num in range(1, XLSX_EXPORT_FIRST_DATA_ROW):
                for col_num in range(1, len(XLSX_EXPORT_FIELDS) + 1):
                    source_cell = base_sheet.cell(row=row_num, column=col_num)
                    target_cell = worksheet.cell(row=row_num, column=col_num)
                    target_cell.value = source_cell.value
                    target_cell._style = copy(source_cell._style)
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.border = copy(source_cell.border)
                        target_cell.alignment = copy(source_cell.alignment)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy(source_cell.protection)

            for merged_range in base_sheet.merged_cells.ranges:
                if merged_range.max_row < XLSX_EXPORT_FIRST_DATA_ROW:
                    worksheet.merge_cells(str(merged_range))

            worksheet.sheet_view.zoomScale = base_sheet.sheet_view.zoomScale
            worksheet.sheet_view.zoomScaleNormal = base_sheet.sheet_view.zoomScaleNormal

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.scale = None
        worksheet.print_options.horizontalCentered = True
        worksheet.print_title_rows = f"{XLSX_EXPORT_HEADER_ROW}:{XLSX_EXPORT_HEADER_ROW}"
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.35
        worksheet.page_margins.bottom = 0.35

        # Keep template column geometry so logo anchor stays aligned.
        worksheet.row_dimensions[XLSX_EXPORT_HEADER_ROW].height = 36

        # Ensure sheet is fresh before writing this chunk.
        for row_index in range(XLSX_EXPORT_FIRST_DATA_ROW, max(worksheet.max_row, XLSX_EXPORT_FIRST_DATA_ROW) + 1):
            for col_index in range(1, len(XLSX_EXPORT_FIELDS) + 1):
                worksheet.cell(row=row_index, column=col_index).value = None

        for column_index, (_field_name, label) in enumerate(XLSX_EXPORT_FIELDS, start=1):
            cell = worksheet.cell(row=XLSX_EXPORT_HEADER_ROW, column=column_index)
            cell.value = label
            cell.font = Font(bold=True, size=XLSX_EXPORT_HEADER_FONT_SIZE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        for row_offset, row in enumerate(chunk, start=0):
            row_index = XLSX_EXPORT_FIRST_DATA_ROW + row_offset
            row_data = _extract_xlsx_row(row)
            worksheet.row_dimensions[row_index].height = XLSX_EXPORT_ROW_HEIGHT
            for column_index, (field_name, _label) in enumerate(XLSX_EXPORT_FIELDS, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.value = row_data[field_name]
                cell.font = Font(size=XLSX_EXPORT_DATA_FONT_SIZE)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = cell_border

        last_data_row = max(XLSX_EXPORT_HEADER_ROW, XLSX_EXPORT_FIRST_DATA_ROW + len(chunk) - 1)
        worksheet.print_area = f"A1:L{last_data_row}"

    return workbook


def _excel_column_letter(column_index: int) -> str:
    letters = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _build_export_ranges(total_records: int) -> list[dict[str, int | str]]:
    ranges: list[dict[str, int | str]] = []
    for start in range(1, total_records + 1, XLSX_EXPORT_ENTRIES_PER_SHEET):
        end = min(start + XLSX_EXPORT_ENTRIES_PER_SHEET - 1, total_records)
        page = ((start - 1) // XLSX_EXPORT_ENTRIES_PER_SHEET) + 1
        ranges.append({"page": page, "start": start, "end": end, "label": f"{start}-{end}"})
    return ranges


@app.context_processor
def inject_export_ranges() -> dict[str, list[dict[str, int | str]]]:
    db = get_db()
    total_records = db.execute("SELECT COUNT(*) AS total FROM pcr_reports").fetchone()["total"]
    return {"export_ranges": _build_export_ranges(total_records)}


def _build_prefill(form_data: dict | None = None, row: sqlite3.Row | None = None) -> dict:
    form_data = form_data or {}

    sample_history = form_data.get("sample_history", {})
    patient_assessment = form_data.get("patient_assessment", {})
    physical_exam = form_data.get("physical_exam", {})
    gcs = form_data.get("gcs", {})
    obstetrics = form_data.get("obstetrics_data", {})
    contact_dispatch = form_data.get("contact_dispatch", {})
    incident_details = form_data.get("incident_details", {})
    team_destination = form_data.get("team_destination", {})
    mvc = form_data.get("mvc", {})
    consent_refusal = form_data.get("consent_refusal", {})
    if row is None:
        patient_name = ""
        nature_of_call = ""
        call_date = ""
        time_of_call = ""
        status = "SUBMITTED"
    else:
        patient_name = row["patient_name"] or ""
        nature_of_call = row["nature_of_call"] or ""
        call_date = row["call_date"] or ""
        time_of_call = row["time_of_call"] or ""
        status = row["status"] or "SUBMITTED"

    return {
        "simple": {
            "patient_name": patient_name,
            "age": form_data.get("patient_info", {}).get("age", ""),
            "gender": form_data.get("patient_info", {}).get("gender", ""),
            "nationality": form_data.get("patient_info", {}).get("nationality", ""),
            "call_date": call_date,
            "time_of_call": time_of_call,
            "status": status,
            "permanent_address": contact_dispatch.get("permanent_address", ""),
            "contact_number": contact_dispatch.get("contact_number", ""),
            "etd_base": contact_dispatch.get("etd_base", ""),
            "eta_scene": contact_dispatch.get("eta_scene", ""),
            "etd_scene": incident_details.get("etd_scene", ""),
            "eta_hospital": incident_details.get("eta_hospital", ""),
            "etd_hospital": incident_details.get("etd_hospital", ""),
            "eta_base": incident_details.get("eta_base", ""),
            "location_of_incident": incident_details.get("location_of_incident", ""),
            "nature_of_illness": incident_details.get("nature_of_illness", ""),
            "mechanism_of_injury": incident_details.get("mechanism_of_injury", ""),
            "chief_complaint": patient_assessment.get("chief_complaint", ""),
            "body_diagram_drawing": physical_exam.get("body_diagram_drawing", ""),
            "deformity": physical_exam.get("deformity", ""),
            "bleeding": physical_exam.get("bleeding", ""),
            "contusion": physical_exam.get("contusion", ""),
            "tenderness": physical_exam.get("tenderness", ""),
            "abrasion": physical_exam.get("abrasion", ""),
            "laceration": physical_exam.get("laceration", ""),
            "punctured": physical_exam.get("punctured", ""),
            "swelling": physical_exam.get("swelling", ""),
            "pupils": patient_assessment.get("pupils", ""),
            "eye_opening": gcs.get("eye_opening", ""),
            "verbal_response": gcs.get("verbal_response", ""),
            "motor_response": gcs.get("motor_response", ""),
            "gcs_total": gcs.get("gcs_total", ""),
            "last_menstrual_period": obstetrics.get("last_menstrual_period", ""),
            "gravida": obstetrics.get("gravida", ""),
            "pre_term": obstetrics.get("pre_term", ""),
            "age_of_gestation": obstetrics.get("age_of_gestation", ""),
            "para": obstetrics.get("para", ""),
            "abortion": obstetrics.get("abortion", ""),
            "expected_date_of_delivery": obstetrics.get("expected_date_of_delivery", ""),
            "term": obstetrics.get("term", ""),
            "living": obstetrics.get("living", ""),
            "onset": form_data.get("opqrst", {}).get("onset", ""),
            "provoking_factors": form_data.get("opqrst", {}).get("provoking_factors", ""),
            "quality": form_data.get("opqrst", {}).get("quality", ""),
            "radiation": form_data.get("opqrst", {}).get("radiation", ""),
            "severity": form_data.get("opqrst", {}).get("severity", ""),
            "timing": form_data.get("opqrst", {}).get("timing", ""),
            "narrative": form_data.get("narrative_report", ""),
            "ambulance_driver": team_destination.get("ambulance_driver", ""),
            "plate_no": team_destination.get("plate_no", ""),
            "license_no": team_destination.get("license_no", ""),
            "responders_tl": team_destination.get("responders_tl", ""),
            "crew": team_destination.get("crew", ""),
            "crew_members": _split_csv_values(team_destination.get("crew", "")),
            "destination_determination": team_destination.get("destination_determination", ""),
            "receiving_facility": team_destination.get("receiving_facility", ""),
            "receiving_personnel": team_destination.get("receiving_personnel", ""),
            "communicator": team_destination.get("communicator", "") or team_destination.get("receiving_personnel", ""),
            "remarks": form_data.get("narrative_report", ""),
            "mvc_type_of_vehicles": mvc.get("type_of_vehicles", ""),
            "mvc_plate_numbers": mvc.get("plate_numbers", ""),
            "mvc_type_of_accident": mvc.get("type_of_accident", ""),
            "law_enforcer_name": mvc.get("law_enforcer_name", ""),
            "law_enforcer_contact": mvc.get("law_enforcer_contact", ""),
            "consent_patient_name": consent_refusal.get("consent_patient_name", ""),
            "consent_date": consent_refusal.get("consent_date", ""),
            "consent_time": consent_refusal.get("consent_time", ""),
            "refusal_treatment_agreement": consent_refusal.get("refusal_treatment_agreement", "Disagreed"),
            "refusal_admission_agreement": consent_refusal.get("refusal_admission_agreement", "Disagreed"),
        },
        "radio": {
            "nature_of_call": nature_of_call,
            "c_spine": patient_assessment.get("c_spine", ""),
            "airway": patient_assessment.get("airway", ""),
            "loc_level": patient_assessment.get("loc_level", ""),
            "capillary_refill": patient_assessment.get("capillary_refill", ""),
        },
        "multi": {
            "types_of_emergencies": contact_dispatch.get("types_of_emergencies", []),
            "breathing": patient_assessment.get("breathing", []),
            "circulation": patient_assessment.get("circulation", []),
            "skin": form_data.get("physical_exam", {}).get("skin", []),
            "burns": form_data.get("physical_exam", {}).get("burns", []),
            "findings": form_data.get("physical_exam", {}).get("findings", []),
            "care_airway": form_data.get("care_management", {}).get("airway", []),
            "care_breathing": form_data.get("care_management", {}).get("breathing", []),
            "care_circulation": form_data.get("care_management", {}).get("circulation", []),
            "care_immobilization": form_data.get("care_management", {}).get("immobilization", []),
            "care_wound": form_data.get("care_management", {}).get("wound_care", []),
        },
        "symptoms": _split_csv_values(sample_history.get("symptoms", "")),
        "next_of_kin": form_data.get("next_of_kin", {}).get("kins", []),
        "vital_signs": form_data.get("vital_signs", []),
    }


def _flatten_for_csv(value, prefix: str, output: dict[str, str]) -> None:
    if isinstance(value, dict):
        for key, nested_value in value.items():
            next_prefix = f"{prefix}_{key}" if prefix else key
            _flatten_for_csv(nested_value, next_prefix, output)
        return

    if isinstance(value, list):
        if not value:
            output[prefix] = "N/A"
            return

        if all(isinstance(item, dict) for item in value):
            for index, item in enumerate(value, start=1):
                _flatten_for_csv(item, f"{prefix}_{index}", output)
        else:
            output[prefix] = " | ".join(str(item) for item in value)
        return

    text = "" if value is None else str(value).strip()
    output[prefix] = text if text else "N/A"


def _extract_xlsx_row(row: sqlite3.Row) -> dict[str, str]:
    form_data = _load_form_data(row)
    contact_dispatch = form_data.get("contact_dispatch", {})
    patient_info = form_data.get("patient_info", {})
    incident_details = form_data.get("incident_details", {})
    team_destination = form_data.get("team_destination", {})
    narrative = form_data.get("narrative_report", "")
    communicator = team_destination.get("communicator", "") or team_destination.get("receiving_personnel", "")

    type_of_emergency = contact_dispatch.get("types_of_emergencies", [])
    if isinstance(type_of_emergency, list):
        type_of_emergency = ", ".join(type_of_emergency)

    return {
        "type_of_emergency": type_of_emergency or "",
        "chief_complaint": form_data.get("patient_assessment", {}).get("chief_complaint", ""),
        "patient_name": row["patient_name"] or "",
        "patient_address": contact_dispatch.get("permanent_address", ""),
        "sex": patient_info.get("gender", ""),
        "date_of_incident": row["call_date"] or "",
        "time_of_incident": row["time_of_call"] or "",
        "place_of_incident": incident_details.get("location_of_incident", ""),
        "driver": team_destination.get("ambulance_driver", ""),
        "responders": team_destination.get("responders_tl", "") or ", ".join(team_destination.get("crew_members", [])) or team_destination.get("crew", ""),
        "communicator": communicator,
        "remarks": narrative,
    }


def _collect_form_data() -> dict:
    return {
        "patient_info": {
            "age": request.form.get("age", "").strip(),
            "gender": request.form.get("gender", "").strip(),
            "nationality": request.form.get("nationality", "").strip(),
            "date": request.form.get("call_date", "").strip(),
            "time_of_call": request.form.get("time_of_call", "").strip(),
        },
        "contact_dispatch": {
            "permanent_address": request.form.get("permanent_address", "").strip(),
            "contact_number": request.form.get("contact_number", "").strip(),
            "etd_base": request.form.get("etd_base", "").strip(),
            "types_of_emergencies": _collect_multi("types_of_emergencies"),
            "eta_scene": request.form.get("eta_scene", "").strip(),
        },
        "next_of_kin": {
            "kins": _collect_kin_entries(),
        },
        "incident_details": {
            "location_of_incident": request.form.get("location_of_incident", "").strip(),
            "nature_of_illness": request.form.get("nature_of_illness", "").strip(),
            "mechanism_of_injury": request.form.get("mechanism_of_injury", "").strip(),
            "etd_scene": request.form.get("etd_scene", "").strip(),
            "eta_hospital": request.form.get("eta_hospital", "").strip(),
            "etd_hospital": request.form.get("etd_hospital", "").strip(),
            "eta_base": request.form.get("eta_base", "").strip(),
        },
        "obstetrics_data": {
            "last_menstrual_period": request.form.get("last_menstrual_period", "").strip(),
            "gravida": request.form.get("gravida", "").strip(),
            "pre_term": request.form.get("pre_term", "").strip(),
            "age_of_gestation": request.form.get("age_of_gestation", "").strip(),
            "para": request.form.get("para", "").strip(),
            "abortion": request.form.get("abortion", "").strip(),
            "expected_date_of_delivery": request.form.get("expected_date_of_delivery", "").strip(),
            "term": request.form.get("term", "").strip(),
            "living": request.form.get("living", "").strip(),
        },
        "patient_assessment": {
            "chief_complaint": request.form.get("chief_complaint", "").strip(),
            "c_spine": request.form.get("c_spine", "").strip(),
            "airway": request.form.get("airway", "").strip(),
            "breathing": _collect_multi("breathing"),
            "circulation": _collect_multi("circulation"),
            "pupils": request.form.get("pupils", "").strip(),
            "loc_level": request.form.get("loc_level", "").strip(),
            "capillary_refill": request.form.get("capillary_refill", "").strip(),
        },
        "gcs": {
            "eye_opening": request.form.get("eye_opening", "").strip(),
            "verbal_response": request.form.get("verbal_response", "").strip(),
            "motor_response": request.form.get("motor_response", "").strip(),
            "gcs_total": request.form.get("gcs_total", "").strip(),
        },
        "vital_signs": _collect_vital_rows(),
        "sample_history": {
            "symptoms": request.form.get("symptoms", "").strip(),
            "allergies": request.form.get("allergies", "").strip(),
            "medications": request.form.get("medications", "").strip(),
            "past_medical_history": request.form.get("past_medical_history", "").strip(),
            "last_oral_intake": request.form.get("last_oral_intake", "").strip(),
            "events_prior": request.form.get("events_prior", "").strip(),
        },
        "opqrst": {
            "onset": request.form.get("onset", "").strip(),
            "provoking_factors": request.form.get("provoking_factors", "").strip(),
            "quality": request.form.get("quality", "").strip(),
            "radiation": request.form.get("radiation", "").strip(),
            "severity": request.form.get("severity", "").strip(),
            "timing": request.form.get("timing", "").strip(),
        },
        "physical_exam": {
            "skin": _collect_multi("skin"),
            "burns": _collect_multi("burns"),
            "findings": _collect_multi("findings"),
            "body_diagram_drawing": request.form.get("body_diagram_drawing", "").strip(),
            "deformity": request.form.get("deformity", "").strip(),
            "bleeding": request.form.get("bleeding", "").strip(),
            "contusion": request.form.get("contusion", "").strip(),
            "tenderness": request.form.get("tenderness", "").strip(),
            "abrasion": request.form.get("abrasion", "").strip(),
            "laceration": request.form.get("laceration", "").strip(),
            "punctured": request.form.get("punctured", "").strip(),
            "swelling": request.form.get("swelling", "").strip(),
        },
        "narrative_report": request.form.get("remarks", request.form.get("narrative", "")).strip(),
        "team_destination": {
            "ambulance_driver": request.form.get("ambulance_driver", "").strip(),
            "plate_no": request.form.get("plate_no", "").strip(),
            "license_no": request.form.get("license_no", "").strip(),
            "responders_tl": request.form.get("responders_tl", "").strip(),
            "crew": ", ".join(_collect_multi("crew_member")) or request.form.get("crew", "").strip(),
            "crew_members": _collect_multi("crew_member"),
            "destination_determination": request.form.get("destination_determination", "").strip(),
            "receiving_facility": request.form.get("receiving_facility", "").strip(),
            "communicator": request.form.get("communicator", "").strip(),
            "receiving_personnel": request.form.get("receiving_personnel", "").strip(),
        },
        "care_management": {
            "airway": _collect_multi("care_airway"),
            "breathing": _collect_multi("care_breathing"),
            "circulation": _collect_multi("care_circulation"),
            "immobilization": _collect_multi("care_immobilization"),
            "wound_care": _collect_multi("care_wound"),
        },
        "mvc": {
            "type_of_vehicles": request.form.get("mvc_type_of_vehicles", "").strip(),
            "plate_numbers": request.form.get("mvc_plate_numbers", "").strip(),
            "type_of_accident": request.form.get("mvc_type_of_accident", "").strip(),
            "law_enforcer_name": request.form.get("law_enforcer_name", "").strip(),
            "law_enforcer_contact": request.form.get("law_enforcer_contact", "").strip(),
        },
        "consent_refusal": {
            "consent_patient_name": request.form.get("consent_patient_name", "").strip(),
            "consent_date": request.form.get("consent_date", "").strip(),
            "consent_time": request.form.get("consent_time", "").strip(),
            "refusal_treatment_agreement": _agreement_status(request.form.getlist("refusal_treatment_agreement")),
            "refusal_admission_agreement": _agreement_status(request.form.getlist("refusal_admission_agreement")),
        },
    }


@app.route("/new", methods=["GET", "POST"])
def new_pcr():
    if request.method == "POST":
        patient_name = request.form.get("patient_name", "").strip()
        nature_of_call = request.form.get("nature_of_call", "").strip()
        call_date = request.form.get("call_date", "").strip()
        time_of_call = request.form.get("time_of_call", "").strip()
        status = request.form.get("status", "SUBMITTED").strip() or "SUBMITTED"

        if not patient_name:
            flash("Patient Name is required.", "error")
            return render_template("new.html")

        form_data = _collect_form_data()

        db = get_db()
        try:
            db.execute(
                """
                INSERT INTO pcr_reports (
                    patient_name, nature_of_call, call_date, time_of_call,
                    status, form_data, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    patient_name,
                    nature_of_call,
                    call_date,
                    time_of_call,
                    status,
                    json.dumps(form_data, ensure_ascii=True),
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            db.commit()
        except sqlite3.Error:
            db.rollback()
            flash("Failed to save PCR report due to a database error.", "error")
            return render_template(
                "new.html",
                form_action=url_for("new_pcr"),
                page_title="PATIENT CARE REPORT",
                submit_label="Save PCR",
                prefill=_build_prefill(form_data),
            )

        flash("PCR report saved.", "success")
        return redirect(url_for("records"))

    return render_template(
        "new.html",
        form_action=url_for("new_pcr"),
        page_title="PATIENT CARE REPORT",
        submit_label="Save PCR",
        prefill=_build_prefill(),
    )


@app.route("/records")
def records():
    search_name = request.args.get("q", "").strip()
    search_date = request.args.get("date", "").strip()

    db = get_db()
    query = "SELECT * FROM pcr_reports WHERE 1=1"
    params: list[str] = []
    if search_name:
        query += " AND patient_name LIKE ?"
        params.append(f"%{search_name}%")
    if search_date:
        query += " AND call_date = ?"
        params.append(search_date)
    query += " ORDER BY id ASC"

    rows = db.execute(query, params).fetchall()
    return render_template(
        "records.html",
        records=rows,
        search_name=search_name,
        search_date=search_date,
    )


@app.route("/records/<int:report_id>")
def view_record(report_id: int):
    db = get_db()
    row = db.execute(
        "SELECT * FROM pcr_reports WHERE id = ?",
        (report_id,),
    ).fetchone()
    if row is None:
        flash("PCR record not found.", "error")
        return redirect(url_for("records"))

    form_data = {}
    form_data = _load_form_data(row)

    return render_template("view.html", record=row, form_data=form_data)


@app.route("/records/<int:report_id>/edit", methods=["GET", "POST"])
def edit_record(report_id: int):
    db = get_db()
    row = db.execute(
        "SELECT * FROM pcr_reports WHERE id = ?",
        (report_id,),
    ).fetchone()
    if row is None:
        flash("PCR record not found.", "error")
        return redirect(url_for("records"))

    form_data = _load_form_data(row)
    prefill = _build_prefill(form_data, row)

    if request.method == "POST":
        patient_name = request.form.get("patient_name", "").strip()
        nature_of_call = request.form.get("nature_of_call", "").strip()
        call_date = request.form.get("call_date", "").strip()
        time_of_call = request.form.get("time_of_call", "").strip()
        status = request.form.get("status", "SUBMITTED").strip() or "SUBMITTED"

        if not patient_name:
            flash("Patient Name is required.", "error")
        else:
            form_data = _collect_form_data()

            try:
                db.execute(
                    """
                    UPDATE pcr_reports
                    SET patient_name = ?, nature_of_call = ?, call_date = ?, time_of_call = ?,
                        status = ?, form_data = ?
                    WHERE id = ?
                    """,
                    (
                        patient_name,
                        nature_of_call,
                        call_date,
                        time_of_call,
                        status,
                        json.dumps(form_data, ensure_ascii=True),
                        report_id,
                    ),
                )
                db.commit()
            except sqlite3.Error:
                db.rollback()
                flash("Failed to update PCR report due to a database error.", "error")
                prefill = _build_prefill(form_data, row)
                return render_template(
                    "new.html",
                    record=row,
                    form_data=form_data,
                    prefill=prefill,
                    form_action=url_for("edit_record", report_id=report_id),
                    page_title=f"EDIT PCR RECORD #{report_id}",
                    submit_label="Update PCR",
                )

            flash("PCR record updated.", "success")
            return redirect(url_for("view_record", report_id=report_id))

    return render_template(
        "new.html",
        record=row,
        form_data=form_data,
        prefill=prefill,
        form_action=url_for("edit_record", report_id=report_id),
        page_title=f"EDIT PCR RECORD #{report_id}",
        submit_label="Update PCR",
    )


@app.route("/records/<int:report_id>/delete", methods=["POST"])
def delete_record(report_id: int):
    db = get_db()
    try:
        db.execute("DELETE FROM pcr_reports WHERE id = ?", (report_id,))
        remaining = db.execute("SELECT COUNT(*) AS total FROM pcr_reports").fetchone()["total"]
        if remaining == 0:
            # Reset autoincrement so a fresh dataset starts again at ID 1.
            db.execute("DELETE FROM sqlite_sequence WHERE name = 'pcr_reports'")
        db.commit()
    except sqlite3.Error:
        db.rollback()
        flash("Failed to delete PCR record due to a database error.", "error")
        return redirect(url_for("records"))
    flash("PCR record deleted.", "success")
    return redirect(url_for("records"))


@app.route("/records/delete-all", methods=["POST"])
def delete_all_records():
    db = get_db()
    try:
        db.execute("DELETE FROM pcr_reports")
        # Reset autoincrement so next record starts at ID 1.
        db.execute("DELETE FROM sqlite_sequence WHERE name = 'pcr_reports'")
        db.commit()
    except sqlite3.Error:
        db.rollback()
        flash("Failed to delete all records due to a database error.", "error")
        return redirect(url_for("records"))

    flash("All PCR records have been deleted.", "success")
    return redirect(url_for("records"))


@app.route("/records/export.csv")
def export_records_csv():
    db = get_db()
    rows = db.execute("SELECT * FROM pcr_reports ORDER BY id ASC").fetchall()

    export_rows: list[dict[str, str]] = []
    fieldnames: list[str] = [
        "id",
        "patient_name",
        "nature_of_call",
        "call_date",
        "time_of_call",
        "status",
        "created_at",
        "form_data_json",
    ]
    extra_fields: set[str] = set()

    for row in rows:
        row_data: dict[str, str] = {
            "id": str(row["id"]),
            "patient_name": row["patient_name"] or "",
            "nature_of_call": row["nature_of_call"] or "",
            "call_date": row["call_date"] or "",
            "time_of_call": row["time_of_call"] or "",
            "status": row["status"] or "",
            "created_at": row["created_at"] or "",
            "form_data_json": row["form_data"] or "{}",
        }

        form_data = _load_form_data(row)
        flattened: dict[str, str] = {}
        _flatten_for_csv(form_data, "form_data", flattened)
        row_data.update(flattened)
        extra_fields.update(flattened.keys())
        export_rows.append(row_data)

    fieldnames.extend(sorted(extra_fields))

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row_data in export_rows:
        writer.writerow(row_data)

    csv_data = output.getvalue()
    output.close()

    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=pcr_records.csv"},
    )


@app.route("/records/export.xlsx")
def export_records_xlsx():
    selected_page_raw = request.args.get("page", "").strip()
    selected_page = None
    if selected_page_raw:
        try:
            selected_page = int(selected_page_raw)
            if selected_page < 1:
                raise ValueError
        except ValueError:
            flash("Invalid XLSX range selection.", "error")
            return redirect(url_for("records"))

    db = get_db()
    if selected_page is None:
        rows = db.execute("SELECT * FROM pcr_reports ORDER BY id ASC").fetchall()
        download_name = "pcr_records.xlsx"
    else:
        offset = (selected_page - 1) * XLSX_EXPORT_ENTRIES_PER_SHEET
        rows = db.execute(
            "SELECT * FROM pcr_reports ORDER BY id ASC LIMIT ? OFFSET ?",
            (XLSX_EXPORT_ENTRIES_PER_SHEET, offset),
        ).fetchall()
        if not rows:
            flash("Selected XLSX range is empty.", "error")
            return redirect(url_for("records"))

        start_record = offset + 1
        end_record = offset + len(rows)
        download_name = f"pcr_records_{start_record}-{end_record}.xlsx"

    workbook = _build_xlsx_workbook(rows)
    output = io.BytesIO()
    workbook.save(output)

    return send_file(
        io.BytesIO(output.getvalue()),
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/records/export.db")
def export_records_db():
    backup_path = _create_db_backup_file()
    return send_file(
        backup_path,
        as_attachment=True,
        download_name=backup_path.name,
        mimetype="application/octet-stream",
    )


if __name__ == "__main__":
    init_db()
    app.run(debug=True)
